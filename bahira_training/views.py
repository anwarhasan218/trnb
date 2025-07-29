from django.shortcuts import render, redirect, get_object_or_404
from .models import Teacher, TrainingProgram, Enrollment
from .forms import TeacherForm
import openpyxl
from django.http import HttpResponse
from django.db import models
from django import forms
from django.contrib import messages
from django.db.models import Count
import io
import openpyxl
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.http import FileResponse
from math import ceil
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
import arabic_reshaper
from bidi.algorithm import get_display
from reportlab.lib.units import cm

# قائمة المعلمين
def teacher_list(request):
    teachers = Teacher.objects.all()
    query = request.GET.get('q', '').strip()
    subject_filter = request.GET.get('subject', '').strip()
    job_grade_filter = request.GET.get('job_grade', '').strip()

    if query:
        teachers = teachers.filter(
            models.Q(name__icontains=query) |
            models.Q(national_id__icontains=query) |
            models.Q(record_number__icontains=query)
        )
    if subject_filter:
        teachers = teachers.filter(subject=subject_filter)
    if job_grade_filter:
        teachers = teachers.filter(job_grade=job_grade_filter)

    # لجلب القيم المميزة للفلاتر
    subjects = Teacher.objects.values_list('subject', flat=True).distinct()
    job_grades = Teacher.objects.values_list('job_grade', flat=True).distinct()

    return render(request, 'teachers/teacher_list.html', {
        'teachers': teachers,
        'subjects': subjects,
        'job_grades': job_grades,
        'query': query,
        'subject_filter': subject_filter,
        'job_grade_filter': job_grade_filter,
    })

# إضافة معلم
def teacher_add(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('teacher_list')
    else:
        form = TeacherForm()
    return render(request, 'teachers/teacher_form.html', {'form': form, 'title': 'إضافة معلم'})

# تعديل معلم
def teacher_edit(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        form = TeacherForm(request.POST, instance=teacher)
        if form.is_valid():
            form.save()
            return redirect('teacher_list')
    else:
        form = TeacherForm(instance=teacher)
    return render(request, 'teachers/teacher_form.html', {'form': form, 'title': 'تعديل معلم'})

# حذف معلم
def teacher_delete(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        teacher.delete()
        return redirect('teacher_list')
    return render(request, 'teachers/teacher_confirm_delete.html', {'teacher': teacher})

# استيراد المعلمين من Excel (سيتم تنفيذها لاحقاً)
def teacher_import(request):
    message = None
    message_type = "info"
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            expected_headers = [
                'اسم المعلم', 'الرقم القومي', 'رقم السجل', 'المادة', 'المرحلة التعليمية',
                'المنطقة', 'الإدارة', 'المعهد', 'الدرجة الوظيفية'
            ]
            if headers != expected_headers:
                message = "تنسيق الأعمدة غير صحيح. يرجى استخدام النموذج المرفق وعدم تغيير ترتيب أو أسماء الأعمدة."
                message_type = "danger"
            else:
                added, skipped = 0, 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    name, national_id, record_number, subject, stage, region, administration, institute, job_grade = row
                    if not (name and national_id and record_number):
                        skipped += 1
                        continue
                    # منع التكرار بناءً على الرقم القومي أو السجل
                    if Teacher.objects.filter(national_id=national_id).exists() or Teacher.objects.filter(record_number=record_number).exists():
                        skipped += 1
                        continue
                    Teacher.objects.create(
                        name=name,
                        national_id=national_id,
                        record_number=record_number,
                        subject=subject or '',
                        stage=stage or '',
                        region=region or '',
                        administration=administration or '',
                        institute=institute or '',
                        job_grade=job_grade or '',
                    )
                    added += 1
                message = f"تم استيراد {added} معلم بنجاح. تم تجاهل {skipped} صف بسبب نقص البيانات أو التكرار."
                message_type = "success" if added else "warning"
        except Exception as e:
            message = f"حدث خطأ أثناء معالجة الملف: {e}"
            message_type = "danger"
    return render(request, 'teachers/teacher_import.html', {'message': message, 'message_type': message_type})

def download_teacher_template(request):
    # إنشاء ملف Excel فارغ مع رؤوس الأعمدة فقط
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المعلمين'
    ws.append([
        'اسم المعلم', 'الرقم القومي', 'رقم السجل', 'المادة', 'المرحلة التعليمية',
        'المنطقة', 'الإدارة', 'المعهد', 'الدرجة الوظيفية'
    ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=teacher_template.xlsx'
    wb.save(response)
    return response

def delete_all_teachers(request):
    if request.method == 'POST':
        Teacher.objects.all().delete()
        return redirect('teacher_list')
    return render(request, 'teachers/teacher_confirm_delete_all.html')

# نموذج برنامج تدريبي
class TrainingProgramForm(forms.ModelForm):
    class Meta:
        model = TrainingProgram
        fields = ['name', 'target_group', 'stage', 'start_date', 'end_date', 'location', 'room_number']
        labels = {
            'name': 'اسم البرنامج التدريبي',
            'target_group': 'الفئة المستهدفة',
            'stage': 'المرحلة التعليمية',
            'start_date': 'تاريخ البداية',
            'end_date': 'تاريخ النهاية',
            'location': 'مكان إقامة البرنامج التدريبي',
            'room_number': 'رقم القاعة',
        }

# قائمة البرامج التدريبية
def program_list(request):
    programs = TrainingProgram.objects.all()
    return render(request, 'programs/program_list.html', {'programs': programs})

# إضافة برنامج تدريبي
def program_add(request):
    if request.method == 'POST':
        form = TrainingProgramForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('program_list')
    else:
        form = TrainingProgramForm()
    return render(request, 'programs/program_form.html', {'form': form, 'title': 'إضافة برنامج تدريبي'})

# تعديل برنامج تدريبي
def program_edit(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    if request.method == 'POST':
        form = TrainingProgramForm(request.POST, instance=program)
        if form.is_valid():
            form.save()
            return redirect('program_list')
    else:
        form = TrainingProgramForm(instance=program)
    return render(request, 'programs/program_form.html', {'form': form, 'title': 'تعديل برنامج تدريبي'})

# حذف برنامج تدريبي
def program_delete(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    if request.method == 'POST':
        program.delete()
        return redirect('program_list')
    return render(request, 'programs/program_confirm_delete.html', {'program': program})

# تفاصيل برنامج تدريبي
def program_detail(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    enrollments = Enrollment.objects.filter(training_program=program).select_related('teacher')
    enrollments_count = enrollments.count()
    message = None
    message_type = "info"
    # بحث عن معلم
    search_query = request.GET.get('q', '').strip()
    search_results = []
    if search_query:
        search_results = Teacher.objects.filter(
            models.Q(name__icontains=search_query) |
            models.Q(national_id__icontains=search_query) |
            models.Q(record_number__icontains=search_query)
        )
    return render(request, 'programs/program_detail.html', {
        'program': program,
        'enrollments': enrollments,
        'enrollments_count': enrollments_count,
        'search_query': search_query,
        'search_results': search_results,
        'message': message,
        'message_type': message_type,
    })

# إضافة معلم للبرنامج التدريبي
def enroll_teacher(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    if request.method == 'POST':
        teacher_id = request.POST.get('teacher_id')
        teacher = get_object_or_404(Teacher, pk=teacher_id)
        # منع التكرار حسب الرقم القومي أو السجل
        already_enrolled = Enrollment.objects.filter(
            training_program=program
        ).filter(
            models.Q(teacher__national_id=teacher.national_id) |
            models.Q(teacher__record_number=teacher.record_number)
        ).exists()
        if already_enrolled:
            messages.warning(request, f"المعلم '{teacher.name}' مسجل بالفعل في هذا التدريب بنفس الرقم القومي أو السجل.")
        else:
            Enrollment.objects.create(teacher=teacher, training_program=program)
            messages.success(request, f"تم إضافة المعلم '{teacher.name}' بنجاح للبرنامج التدريبي.")
    return redirect('program_detail', pk=pk)

# إزالة معلم من البرنامج التدريبي
def unenroll_teacher(request, pk, teacher_id):
    program = get_object_or_404(TrainingProgram, pk=pk)
    teacher = get_object_or_404(Teacher, pk=teacher_id)
    Enrollment.objects.filter(teacher=teacher, training_program=program).delete()
    return redirect('program_detail', pk=pk)

def dashboard_view(request):
    # عدد البرامج التدريبية الإجمالي
    total_programs = TrainingProgram.objects.count()
    # عدد البرامج لكل تخصص (stage)
    programs_per_stage = TrainingProgram.objects.values('stage').annotate(count=Count('id'))
    # عدد المتدربين الإجمالي (عدد الانتسابات الفريدة)
    total_trainees = Enrollment.objects.values('teacher').distinct().count()
    # عدد المعلمين الإجمالي
    total_teachers = Teacher.objects.count()
    # عدد المتدربين لكل مادة
    trainees_per_subject = Teacher.objects.values('subject').annotate(count=Count('id'))
    # عدد المتدربين لكل مرحلة تعليمية
    trainees_per_stage = Teacher.objects.values('stage').annotate(count=Count('id'))
    return render(request, 'dashboard.html', {
        'total_programs': total_programs,
        'programs_per_stage': programs_per_stage,
        'total_trainees': total_trainees,
        'total_teachers': total_teachers,
        'trainees_per_subject': trainees_per_subject,
        'trainees_per_stage': trainees_per_stage,
    })

def report_view(request):
    from django.db.models import Count
    # جلب الفلاتر من الطلب
    stage_filter = request.GET.get('stage', '').strip()
    subject_filter = request.GET.get('subject', '').strip()
    # جميع البرامج والمعلمين
    programs = TrainingProgram.objects.all()
    teachers = Teacher.objects.all()
    enrollments = Enrollment.objects.all()
    # تطبيق الفلاتر
    if stage_filter:
        programs = programs.filter(stage=stage_filter)
        teachers = teachers.filter(stage=stage_filter)
        enrollments = enrollments.filter(training_program__stage=stage_filter)
    if subject_filter:
        teachers = teachers.filter(subject=subject_filter)
        enrollments = enrollments.filter(teacher__subject=subject_filter)
    # الإحصائيات
    total_programs = programs.count()
    programs_per_stage = programs.values('stage').annotate(count=Count('id'))
    total_trainees = enrollments.values('teacher').distinct().count()
    trainees_per_subject = teachers.values('subject').annotate(count=Count('id'))
    trainees_per_stage = teachers.values('stage').annotate(count=Count('id'))
    # القيم المميزة للفلاتر
    all_stages = TrainingProgram.objects.values_list('stage', flat=True).distinct()
    all_subjects = Teacher.objects.values_list('subject', flat=True).distinct()
    return render(request, 'reports/report_dashboard.html', {
        'total_programs': total_programs,
        'programs_per_stage': programs_per_stage,
        'total_trainees': total_trainees,
        'trainees_per_subject': trainees_per_subject,
        'trainees_per_stage': trainees_per_stage,
        'all_stages': all_stages,
        'all_subjects': all_subjects,
        'stage_filter': stage_filter,
        'subject_filter': subject_filter,
    })

def export_report_excel(request):
    from django.db.models import Count
    stage_filter = request.GET.get('stage', '').strip()
    subject_filter = request.GET.get('subject', '').strip()
    programs = TrainingProgram.objects.all()
    teachers = Teacher.objects.all()
    enrollments = Enrollment.objects.all()
    if stage_filter:
        programs = programs.filter(stage=stage_filter)
        teachers = teachers.filter(stage=stage_filter)
        enrollments = enrollments.filter(training_program__stage=stage_filter)
    if subject_filter:
        teachers = teachers.filter(subject=subject_filter)
        enrollments = enrollments.filter(teacher__subject=subject_filter)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'الإحصائيات'
    ws.append(['العنصر', 'القيمة'])
    total_programs = programs.count()
    total_trainees = enrollments.values('teacher').distinct().count()
    ws.append(['عدد البرامج التدريبية', total_programs])
    ws.append(['عدد المتدربين', total_trainees])
    ws.append([])
    ws.append(['عدد البرامج لكل مرحلة'])
    programs_per_stage = programs.values('stage').annotate(count=Count('id'))
    for item in programs_per_stage:
        ws.append([item['stage'] or 'غير محدد', item['count']])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=report_stats.xlsx'
    return response

def export_report_pdf(request):
    from django.db.models import Count
    stage_filter = request.GET.get('stage', '').strip()
    subject_filter = request.GET.get('subject', '').strip()
    programs = TrainingProgram.objects.all()
    teachers = Teacher.objects.all()
    enrollments = Enrollment.objects.all()
    if stage_filter:
        programs = programs.filter(stage=stage_filter)
        teachers = teachers.filter(stage=stage_filter)
        enrollments = enrollments.filter(training_program__stage=stage_filter)
    if subject_filter:
        teachers = teachers.filter(subject=subject_filter)
        enrollments = enrollments.filter(teacher__subject=subject_filter)
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 50
    p.setFont("Helvetica-Bold", 16)
    p.drawString(200, y, "التقارير والإحصائيات")
    y -= 40
    p.setFont("Helvetica", 12)
    total_programs = programs.count()
    total_trainees = enrollments.values('teacher').distinct().count()
    p.drawString(50, y, f"عدد البرامج التدريبية: {total_programs}")
    y -= 25
    p.drawString(50, y, f"عدد المتدربين: {total_trainees}")
    y -= 35
    p.setFont("Helvetica-Bold", 13)
    p.drawString(50, y, "عدد البرامج لكل مرحلة:")
    y -= 25
    p.setFont("Helvetica", 12)
    programs_per_stage = programs.values('stage').annotate(count=Count('id'))
    for item in programs_per_stage:
        p.drawString(70, y, f"{item['stage'] or 'غير محدد'}: {item['count']}")
        y -= 20
        if y < 60:
            p.showPage()
            y = height - 50
    p.showPage()
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=report_stats.pdf'
    return response

def attendance_sheet_print(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    enrollments = Enrollment.objects.filter(training_program=program).select_related('teacher').order_by('teacher__name')
    teachers = [e.teacher for e in enrollments]
    # تقسيم المعلمين إلى صفحات (20 في كل صفحة)
    pages = [teachers[i:i+20] for i in range(0, len(teachers), 20)]
    days = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
    empty_rows_list = [list(range(20 - len(page))) for page in pages]
    return render(request, 'programs/attendance_sheet_print.html', {
        'program': program,
        'pages': zip(pages, empty_rows_list),
        'days': days,
    })

def introduction_sheet_print(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    enrollments = Enrollment.objects.filter(training_program=program).select_related('teacher').order_by('teacher__name')
    teachers = [e.teacher for e in enrollments]
    # تقسيم المعلمين إلى صفحات (20 في كل صفحة)
    pages = []
    for i, page in enumerate([teachers[i:i+20] for i in range(0, len(teachers), 20)]):
        empty_rows = list(range(20 - len(page)))
        page_start = i * 20
        pages.append((page, empty_rows, page_start))
    return render(request, 'programs/introduction_sheet_print.html', {
        'program': program,
        'pages': pages,
    })

def pretest_sheet_print(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    enrollments = Enrollment.objects.filter(training_program=program).select_related('teacher').order_by('teacher__name')
    teachers = [e.teacher for e in enrollments]
    # إضافة الدرجات والملاحظات إذا كانت موجودة
    for t in teachers:
        t.pretest_score = getattr(t, 'pretest_score', '')
        t.pretest_note = getattr(t, 'pretest_note', '')
    empty_rows = []  # لا حاجة لصفوف فارغة غالباً، أو يمكن ضبطها حسب الحاجة
    pages = [(teachers, empty_rows)]
    return render(request, 'programs/pretest_sheet_print.html', {
        'program': program,
        'pages': pages,
    })

def posttest_sheet_print(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    enrollments = Enrollment.objects.filter(training_program=program).select_related('teacher').order_by('teacher__name')
    teachers = [e.teacher for e in enrollments]
    for t in teachers:
        t.posttest_score = getattr(t, 'posttest_score', '')
        t.posttest_note = getattr(t, 'posttest_note', '')
    empty_rows = []
    pages = [(teachers, empty_rows)]
    return render(request, 'programs/posttest_sheet_print.html', {
        'program': program,
        'pages': pages,
    })

def behavior_sheet_print(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    enrollments = Enrollment.objects.filter(training_program=program).select_related('teacher').order_by('teacher__name')
    teachers = [e.teacher for e in enrollments]
    for t in teachers:
        t.behavior_score = getattr(t, 'behavior_score', '')
        t.commitment_score = getattr(t, 'commitment_score', '')
        t.behavior_note = getattr(t, 'behavior_note', '')
    empty_rows = []
    pages = [(teachers, empty_rows)]
    return render(request, 'programs/behavior_sheet_print.html', {
        'program': program,
        'pages': pages,
    })

def call_sheet_print(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    enrollments = Enrollment.objects.filter(training_program=program).select_related('teacher').order_by('teacher__name')
    trainees = [e.teacher for e in enrollments]
    total_rows = 30
    empty_rows = range(total_rows - len(trainees)) if len(trainees) < total_rows else []
    return render(request, 'programs/call_sheet_print.html', {
        'program': program,
        'trainees': trainees,
        'empty_rows': empty_rows,
    })

def training_completion_notice_print(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    enrollments = Enrollment.objects.filter(training_program=program).select_related('teacher').order_by('teacher__name')
    trainees = [e.teacher for e in enrollments]
    return render(request, 'programs/training_completion_notice_print.html', {
        'program': program,
        'trainees': trainees,
    })

def training_completion_notice_pdf(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    enrollments = Enrollment.objects.filter(training_program=program).select_related('teacher').order_by('teacher__name')
    trainees = [e.teacher for e in enrollments]
    buffer = io.BytesIO()
    from reportlab.lib.pagesizes import landscape
    p = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)
    # تسجيل خط Amiri
    font_path = 'static/fonts/Amiri-Regular.ttf'
    pdfmetrics.registerFont(TTFont('Amiri', font_path))
    # تسجيل خط Amiri-Bold
    font_bold_path = 'static/fonts/Amiri-Bold.ttf'
    pdfmetrics.registerFont(TTFont('Amiri-Bold', font_bold_path))
    def ar(text):
        return get_display(arabic_reshaper.reshape(str(text)))
    for trainee in trainees:
        margin_x = 1 * cm
        margin_y = 0.5 * cm
        y = height - margin_y
        p.setFont("Amiri-Bold", 16)
        p.drawRightString(width - margin_x, y, ar("الأزهر الشريف"))
        y -= 36
        p.setFont("Amiri-Bold", 16)
        p.drawRightString(width - margin_x, y, ar("منطقة البحيرة الأزهرية"))
        y -= 32
        p.setFont("Amiri-Bold", 16)
        p.drawRightString(width - margin_x, y, ar("ادارة التدريب وتنمية المهارات"))
        y -= 40
        p.setFont("Amiri-Bold", 16)
        p.drawCentredString(width/2, y, ar("إخطار إنهاء برنامج تدريبي"))
        y -= 40
        p.setFont("Amiri-Bold", 16)
        p.drawRightString(width - margin_x, y, ar("تفيد إدارة التدريب بمنطقة البحيرة الأزهرية"))
        y -= 40
        p.drawRightString(width - margin_x, y, ar(f"بأن السيد/ {trainee.name} سجل ({trainee.record_number})"))
        y -= 40
        institute_name = trainee.institute if trainee.institute else '---'
        print(f"معهد المعلم: {institute_name}")  # للتشخيص فقط
        p.setFont("Amiri-Bold", 16)
        p.drawRightString(width - margin_x, y, ar(f"بمعهد/ {institute_name}"))
        y -= 40
        p.setFont("Amiri-Bold", 16)
        p.drawRightString(width - margin_x, y, ar(f"قد اجتاز البرنامج التدريبي لمعلمي ({program.target_group}) ({program.name})"))
        y -= 40
        p.drawRightString(width - margin_x, y, ar(f"والمقام : ({program.location or '-'})"))
        y -= 40
        p.drawRightString(width - margin_x, y, ar(f"خلال الفترة من ({program.start_date}) إلى ({program.end_date})"))
        y -= 40
        p.drawRightString(width - margin_x, y, ar("وقد تغيب عن التدريب ايام : ...................................................."))
        y -= 40
        p.drawRightString(width - margin_x, y, ar("وتنبه عليه العودة على مقر عمله صباح اليوم التالي للتدريب."))
        y -= 40
        p.drawRightString(width/2, y, ar("وهذه إفادة منا بذلك ,,,,"))
        # التذييل
        footer_y = 0.5 * cm
        #p.setLineWidth(0.7)
        #p.line(margin_x, footer_y + 1.2*cm, width - margin_x, footer_y + 1.2*cm)
        p.setFont("Amiri-Bold", 15)
        p.drawRightString(width - margin_x, footer_y + 1.5*cm, ar("المختص"))
        p.drawString(margin_x, footer_y + 1.5*cm, ar("يعتمد مدير إدارة التدريب"))
        p.setFont("Amiri-Bold", 15)
        p.drawString(margin_x, footer_y + 0.7*cm, ar("د محمد فتحي قنطوش"))
        p.showPage()
    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='training_completion_notice.pdf')

def attendance_sheet_pdf(request, pk):
    def ar(text):
        import arabic_reshaper
        from bidi.algorithm import get_display
        return get_display(arabic_reshaper.reshape(str(text)))
    program = get_object_or_404(TrainingProgram, pk=pk)
    enrollments = Enrollment.objects.filter(training_program=program).select_related('teacher').order_by('teacher__name')
    teachers = [e.teacher for e in enrollments]
    days = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
    buffer = io.BytesIO()
    font_path = os.path.join('static', 'fonts', 'Amiri-Regular.ttf')
    pdfmetrics.registerFont(TTFont('Amiri', font_path))
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='AmiriTitle', fontName='Amiri', fontSize=14, alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='AmiriNormal', fontName='Amiri', fontSize=13, alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='AmiriTable', fontName='Amiri', fontSize=11, alignment=TA_CENTER))
    num_pages = ceil(len(teachers)/20) or 1
    for page_num in range(num_pages):
        page_teachers = teachers[page_num*20:(page_num+1)*20]
        # رأس الصفحة (كل سطر Paragraph منفصل)
        elements.append(Paragraph(ar("الأزهر الشريف"), styles['AmiriTitle']))
        elements.append(Paragraph(ar("منطقة البحيرة الأزهرية"), styles['AmiriTitle']))
        elements.append(Paragraph(ar("الإدارة العامة للتدريب وتنمية المهارات"), styles['AmiriTitle']))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(ar(f"حضور وانصراف المتدربين لبرنامج ({program.name}) للفئة ({program.target_group})"), styles['AmiriNormal']))
        elements.append(Paragraph(ar(f"المقام في الفترة من {program.start_date} إلى {program.end_date} ({program.location or ''})"), styles['AmiriNormal']))
        elements.append(Spacer(1, 12))
        # الجدول
        table_data = [[ar("م"), ar("الاسم الرباعي")]]
        for day in days:
            table_data[0].extend([ar(f"حضور {day}"), ar(f"انصراف {day}")])
        for idx, teacher in enumerate(page_teachers, start=1+page_num*20):
            row = [ar(idx), ar(teacher.name)]
            for _ in days:
                row.extend(["", ""])
            table_data.append(row)
        for _ in range(20 - len(page_teachers)):
            row = ["", ""] + ["", ""]*len(days)
            table_data.append(row)
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Amiri'),
            ('FONTSIZE', (0,0), (-1,0), 11),
            ('FONTSIZE', (0,1), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('ALIGN', (0,1), (1,-1), 'RIGHT'),
            ('ALIGN', (2,1), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 18))
        # التذييل
        footer = Paragraph(ar("المشرف المنفذ: ____________ التوقيع: ____________ مدير إدارة التدريب: د. محمد فتحي قطبون"), styles['AmiriNormal'])
        elements.append(footer)
        if page_num < num_pages-1:
            elements.append(Spacer(1, 30))
            elements.append(PageBreak())
    doc.build(elements)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='attendance_sheet.pdf')

def attendance_sheet_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side
    import shutil
    program = get_object_or_404(TrainingProgram, pk=pk)
    enrollments = Enrollment.objects.filter(training_program=program).select_related('teacher').order_by('teacher__name')
    teachers = [e.teacher for e in enrollments]
    # فتح القالب
    template_path = 'templates/حضور وانصراف.xlsx'
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    # تعبئة بيانات البرنامج في الخلايا المناسبة (تعديل حسب القالب)
    # مثال: إذا كان اسم البرنامج في الخلية B2، الفئة في B3، التاريخ في B4، المكان في B5
    # عدل الخلايا حسب ما هو موجود فعليًا في القالب
    ws['B2'] = program.name
    ws['B3'] = program.target_group
    ws['B4'] = f"{program.start_date} إلى {program.end_date}"
    ws['B5'] = program.location or '-'
    # تعبئة أسماء المتدربين في الجدول (حدد أول صف للأسماء)
    start_row = 8  # عدل هذا الرقم حسب أول صف للأسماء في القالب
    for idx, teacher in enumerate(teachers, start=1):
        ws.cell(row=start_row+idx-1, column=1).value = idx
        ws.cell(row=start_row+idx-1, column=2).value = teacher.name
    # إعداد الملف
    from django.http import HttpResponse
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=attendance_sheet.xlsx'
    return response

def attendance_sheet_download_options(request, pk):
    program = get_object_or_404(TrainingProgram, pk=pk)
    return render(request, 'programs/attendance_sheet_download_options.html', {'program': program})
